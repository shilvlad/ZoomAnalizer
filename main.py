import openpyxl
import os
import datetime

DIRECTORY = "work_xls"
files = os.listdir(DIRECTORY)

result = openpyxl.Workbook()
result.create_sheet(title = 'Свод', index = 0)
wr_sheet = result.active

result1 = openpyxl.Workbook()
result1.create_sheet(title = 'Свод', index = 0)
wr_sheet1 = result1.active

statistic = []

i=1
for f in files:
    print(f)
    wb = openpyxl.load_workbook(DIRECTORY+"/"+f, data_only=True)
    ws = wb['1']
    for row in ws.iter_rows():
        conf_start = row[7].value
        conf_start_obj = datetime.datetime.strptime(conf_start, '%b %d, %Y %I:%M %p')

        conf_duration = row[9].value
        conf_duration_obj = datetime.datetime.strptime(conf_duration, '%H:%M:%S')
        conf_end_obj = conf_start_obj + datetime.timedelta(hours=conf_duration_obj.hour, minutes=conf_duration_obj.minute)

        conf_population = row[10].value

        wr_sheet.cell(row=i, column=1).value = conf_start_obj.strftime('%Y-%m-%d %H:%M')
        wr_sheet.cell(row=i, column=2).value = conf_duration_obj.strftime('%H:%M')
        wr_sheet.cell(row=i, column=3).value = conf_end_obj.strftime( '%Y-%m-%d %H:%M')
        wr_sheet.cell(row=i, column=4).value = conf_population
        statistic.append((conf_start_obj, conf_end_obj, conf_population))
        i+=1



scale = []
scale_current = datetime.datetime.strptime('2021-01-01 00:00', '%Y-%m-%d %H:%M')
scale_delta_min= 1
j =1
while scale_current < datetime.datetime.now():
    #print(scale_current)
    population = 0
    for i in statistic:
        if scale_current > i[0] and scale_current < i[1]:
            population += i[2]

    wr_sheet1.cell(row=j, column=1).value = scale_current
    wr_sheet1.cell(row=j, column=2).value = population
    j+=1

    scale_current += datetime.timedelta(minutes=1)







if os.path.isfile('result.xlsx'):
    os.remove('result.xlsx')

result.save('result.xlsx')

if os.path.isfile('result1.xlsx'):
    os.remove('result1.xlsx')

result1.save('result1.xlsx')